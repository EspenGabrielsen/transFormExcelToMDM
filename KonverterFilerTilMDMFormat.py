import openpyxl
import warnings
from tkinter import filedialog, messagebox
import json
import os
import datetime 

# Deaktiver advarsler fra openpyxl som alltid dukker opp pga. macro i excel-innlesning?
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

class transformExcel:
    def getInputData(self):
        self.RowsConfig = ""

        try:
            with open("configFK.json", "r") as f:
                try:
                    self.RowsConfig = json.load(f)
                except ValueError as e:
                    print(e, "\nConfig filen virker å være korrupt")
        except FileNotFoundError as e:
            print(e, "\nKunne ikke finne config filen")

        self.filename = filedialog.askopenfilename(
            filetypes=[("Excel-filer", "*.xlsx")], title="Velg fil"
        )

    def variableSetup(self):
        if self.RowsConfig and self.filename:
            self.StartRow = self.RowsConfig.get("startrow") or 1
            self.outPutRow = self.RowsConfig.get("outPutRow") or 1
            self.PrimaryKeyColumn = self.RowsConfig.get("PrimarKeyColumn") or "A"

            self.TargetRowOffset = range(len(self.RowsConfig.get("Rows"))) or 1
            self.itemOfset = len(self.RowsConfig.get("Rows")) or 1

            self.wb = openpyxl.load_workbook(filename=self.filename)
            sheet = self.RowsConfig.get("readSheet") or self.wb.active.title
            try:
                self.ws = self.wb[sheet]
            except:
                self.ws = None

    def setupNewFile(self):
        # Last inn malen for hvor data skal legges inn
        template_file = filedialog.askopenfilename(
            title="Velg malfil der varene skal leses inn",
            filetypes=[("Excel-filer", "*.xlsm;*.xlsx")]
        )

        self.new_wb = openpyxl.load_workbook(template_file, keep_vba=True)  # Behold makroer

        # Sørg for at kun "Entities"-arket er til stede i innlesningsmalen.
        if "Entities" in self.new_wb.sheetnames:
            self.new_ws = self.new_wb["Entities"]
        else:
            messagebox.showwarning("Ark mangler", message="Entities-arket ble ikke funnet i malen.")
            return

        # Fjern alle andre ark enn "Help" og "Entities"
        for sheet in self.new_wb.sheetnames:
            if sheet not in ["Help", "Entities"]:
                del self.new_wb[sheet]  # Fjern uønskede ark

    def readAndWriteExcel(self):
        # Start med å sette gjeldende rad til output-rad
        self.NewWorkBookCurrentRow = self.outPutRow
        self.deleteIfRowFails = []
        for i, self.rowSelected in enumerate(self.RowsConfig["Rows"]):
            for row in range(10_000):
                
                # Beregn den nye raden i output-arbeidsboken basert på offset
                offset = (self.itemOfset * row) + self.skippedRows

                self.NewWorkBookCurrentRow = self.outPutRow + self.TargetRowOffset[i] + offset

                # Stopp hvis PrimaryKeyColumn er tom
                if not self.ws[self.PrimaryKeyColumn + str(self.StartRow + row)].value:
                    break  # Stopp hvis PrimaryKeyColumn er tom

                # Skriv data til den nye arbeidsboken
                for col in self.rowSelected.keys():
                    valueToPlace = (
                        self.rowSelected[col].get("Value")
                        or self.ws[col + str(self.StartRow + row)].value
                    )
                    self.skipRowIfMissingFlagg = True if self.rowSelected[col].get("skipRowIfMissing") and not valueToPlace else False

                    if valueToPlace:
                        target_columns = self.rowSelected[col].get("targetColumn")
                        
                        # Hvis targetColumn er en liste, iterer over den
                        if isinstance(target_columns, list):
                            for target_col in target_columns:
                                self.new_ws[target_col + str(self.NewWorkBookCurrentRow)] = valueToPlace
                        else:

                            self.new_ws[target_columns + str(self.NewWorkBookCurrentRow)] = valueToPlace
                    if self.skipRowIfMissingFlagg:
                        self.deleteIfRowFails.append(self.NewWorkBookCurrentRow)

        self.deleteIfRowFails.sort()
        while self.deleteIfRowFails:
            deleteRow = self.deleteIfRowFails.pop()
            print(f"deleting row {deleteRow}")
            self.new_ws.delete_rows(deleteRow)


    def saveFile(self):
        """Lagre filen med dato og tid."""
        now = datetime.datetime.now()
        date_str = now.strftime("%d%m%H%M")  # Format: ddmmyyhhmm

        # Mappe for å lagre filen som er klar for innlesning
        save_path = filedialog.asksaveasfilename(
            title="Save file as",
            filetypes=[("Excel filer", "*.xlsm")],
            defaultextension=".xlsm",
        )

        # Filnavn som settes ved lagring av filen, lagres som .xlsm pga. makroer i innlesningsfil 
        #save_path = os.path.join(target_folder, f"Innlesing_FK{date_str}.xlsm")

        try:
            self.new_wb.save(save_path)  # Lagre filen på ønsket sted
            messagebox.showinfo("Fil Lagret", f"Filen er lagret som: {save_path}")
            os.system(f'start EXCEL.EXE "{save_path}"')  # Åpne filen i Excel ved lagring for å se
        except PermissionError:
            messagebox.showwarning(title="Fil er låst", message="Filen er allerede åpen. Lukk den og prøv igjen.")

    def __init__(self):
        self.errorColor = openpyxl.styles.fills.PatternFill(
            patternType="solid", fgColor=openpyxl.styles.colors.Color(rgb="00FF0000")
        )
        self.errors = []
        self.getInputData()
        self.variableSetup()
        self.deleteIfRowFails = ""
        self.skippedRows = 0
        if self.ws:
            self.setupNewFile()
            self.readAndWriteExcel()
            self.saveFile()  # Lagre filen
        else:
            messagebox.showwarning("Arbeidsark ikke funnet", message="Kunne ikke finne arbeidsarket")

if __name__ == "__main__":
    transformExcel()
